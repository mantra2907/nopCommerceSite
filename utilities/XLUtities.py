import openpyxl

def getRows(fileName,sheetName):
    workwook=openpyxl.load_workbook(fileName)
    sheet=workwook[sheetName]
    return (sheet.max_row)

def getColumn(fileNmae,sheetName):
    workbook=openpyxl.load_workbook(fileNmae)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def readData(fileName,sheetName,rownum,columnNum):
    workbook=openpyxl.load_workbook(fileName)
    sheet=workbook[sheetName]
    return sheet.cell(row=rownum, column=columnNum).value

def writeData(fileName,sheetName,rownum,columnNum,data):
    workbook=openpyxl.load_workbook(fileName)
    sheet=workbook[sheetName]
    sheet.cell(row=rownum, column=columnNum).value=data
    workbook.save(fileName)






































